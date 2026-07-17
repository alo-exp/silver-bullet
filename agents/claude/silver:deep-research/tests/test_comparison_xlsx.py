"""Tests for comparison-matrix.xlsx export and need-profile personas."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest

SKILL_ROOT = os.path.join(os.path.dirname(__file__), "..")
SCRIPTS = os.path.join(SKILL_ROOT, "scripts")
sys.path.insert(0, SCRIPTS)

from matrix_core import TICK  # noqa: E402
from need_profile_personas import apply_persona_priority, get_persona, list_persona_ids  # noqa: E402


def run_py(script: str, *args: str) -> tuple[int, dict]:
    result = subprocess.run(
        [sys.executable, script, *args],
        capture_output=True,
        text=True,
    )
    out = result.stdout.strip()
    data = json.loads(out) if out.startswith("{") else {}
    return result.returncode, data


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


class TestNeedProfilePersonas(unittest.TestCase):
    def test_startup_persona_exists(self) -> None:
        self.assertIn("startup", list_persona_ids())
        persona = get_persona("startup")
        self.assertIsNotNone(persona)
        assert persona is not None
        self.assertIn("priority_hints", persona)

    def test_persona_raises_sso_priority(self) -> None:
        need = {"persona_id": "enterprise"}
        prio = apply_persona_priority("SSO", "Identity and access", "Medium", need)
        self.assertEqual(prio, "Critical")


class TestGenerateComparisonXlsx(unittest.TestCase):
    def _seed_compare_fixture(self, d: str, *, persona_id: str | None = None) -> None:
        need = {
            "category": "IDP",
            "audience": "CTO",
            "must_haves": ["SSO"],
            "nice_to_haves": [],
            "license_preference": "mixed",
            "interview_complete": True,
        }
        if persona_id:
            need["persona_id"] = persona_id
        with open(os.path.join(d, "need_profile.json"), "w", encoding="utf-8") as f:
            json.dump(need, f)
        for name in ("alpha", "beta"):
            sol = os.path.join(d, "solutions", name)
            os.makedirs(sol)
            with open(os.path.join(sol, "features.json"), "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "solution_name": name,
                        "categories": [
                            {
                                "name": "Core",
                                "features": [
                                    {"name": "SSO", "supported": name == "alpha"},
                                    {"name": "API", "supported": True},
                                ],
                            }
                        ],
                    },
                    f,
                )

    @unittest.skipUnless(_has_openpyxl(), "openpyxl not installed")
    def test_compare_solutions_emits_xlsx_by_default(self) -> None:
        compare = os.path.join(SCRIPTS, "compare_solutions.py")
        with tempfile.TemporaryDirectory() as d:
            self._seed_compare_fixture(d)
            code, data = run_py(compare, "--dir", d)
            self.assertEqual(code, 0)
            self.assertEqual(data["status"], "ok")
            xlsx = os.path.join(d, "comparison", "comparison-matrix.xlsx")
            self.assertTrue(os.path.isfile(xlsx))
            self.assertIn("xlsx", data)

    @unittest.skipUnless(_has_openpyxl(), "openpyxl not installed")
    def test_generate_comparison_xlsx_from_existing_json(self) -> None:
        gen = os.path.join(SCRIPTS, "generate_comparison_xlsx.py")
        with tempfile.TemporaryDirectory() as d:
            self._seed_compare_fixture(d, persona_id="startup")
            compare = os.path.join(SCRIPTS, "compare_solutions.py")
            run_py(compare, "--dir", d, "--no-emit-xlsx")
            code, data = run_py(gen, "--dir", d)
            self.assertEqual(code, 0)
            self.assertEqual(data["status"], "ok")
            xlsx = data["output"]
            self.assertTrue(os.path.isfile(xlsx))

            import openpyxl

            wb = openpyxl.load_workbook(xlsx)
            ws = wb.active
            self.assertIn("startup", str(ws.cell(1, 1).value))
            self.assertEqual(ws.cell(2, 1).value, "Capability / Feature")
            self.assertEqual(ws.cell(2, 3).value, "alpha")
            self.assertEqual(ws.cell(5, 1).value, "Core")
            # SSO row: alpha tick, beta empty
            sso_row = None
            for row in range(5, ws.max_row + 1):
                if ws.cell(row, 1).value == "SSO":
                    sso_row = row
                    break
            self.assertIsNotNone(sso_row)
            assert sso_row is not None
            self.assertEqual(ws.cell(sso_row, 3).value, TICK)
            self.assertIsNone(ws.cell(sso_row, 4).value)
            score_formula = str(ws.cell(4, 3).value or "")
            self.assertTrue(score_formula.startswith("="))
            self.assertIn("COUNTIFS", score_formula)


class TestMatrixBuilderSanitization(unittest.TestCase):
    def test_user_strings_neutralize_formula_prefixes(self) -> None:
        from matrix_builder import _sanitize_cell_value

        self.assertEqual(_sanitize_cell_value("=SUM(A1)"), "'=SUM(A1)")
        self.assertEqual(_sanitize_cell_value("+cmd"), "'+cmd")
        self.assertEqual(_sanitize_cell_value("-1"), "'-1")
        self.assertEqual(_sanitize_cell_value("@evil"), "'@evil")
        self.assertEqual(_sanitize_cell_value("\t=SUM(A1)"), "'\t=SUM(A1)")
        self.assertEqual(_sanitize_cell_value("\r=1"), "'\r=1")
        self.assertEqual(_sanitize_cell_value(" Safe Name"), " Safe Name")


if __name__ == "__main__":
    unittest.main()
