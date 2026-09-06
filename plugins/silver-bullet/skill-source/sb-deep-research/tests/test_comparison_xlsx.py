"""Tests for comparison-matrix.xlsx export and need-profile personas."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

SKILL_ROOT = os.path.join(os.path.dirname(__file__), "..")
SCRIPTS = os.path.join(SKILL_ROOT, "scripts")
sys.path.insert(0, SCRIPTS)

from matrix_core import TICK, comparison_matrix_xlsx_path, research_dir_slug, score_solutions_from_rows  # noqa: E402
from need_profile_personas import apply_persona_priority, get_persona, list_persona_ids  # noqa: E402


def run_py(script: str, *args: str) -> tuple[int, dict]:
    result = subprocess.run(
        [sys.executable, script, *args],
        capture_output=True,
        text=True,
    )
    out = result.stdout.strip()
    data = json.loads(out) if out.startswith("{") else {}
    return result.returncode, data


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


class TestNeedProfilePersonas(unittest.TestCase):
    def test_startup_persona_exists(self) -> None:
        self.assertIn("startup", list_persona_ids())
        persona = get_persona("startup")
        self.assertIsNotNone(persona)
        assert persona is not None
        self.assertIn("priority_hints", persona)

    def test_persona_raises_sso_priority(self) -> None:
        need = {"persona_id": "enterprise"}
        prio = apply_persona_priority("SSO", "Identity and access", "Medium", need)
        self.assertEqual(prio, "Critical")


class TestGenerateComparisonXlsx(unittest.TestCase):
    def _seed_compare_fixture(self, d: str, *, persona_id: str | None = None) -> None:
        need = {
            "category": "IDP",
            "audience": "CTO",
            "must_haves": ["SSO"],
            "nice_to_haves": [],
            "license_preference": "mixed",
            "interview_complete": True,
        }
        if persona_id:
            need["persona_id"] = persona_id
        with open(os.path.join(d, "need_profile.json"), "w", encoding="utf-8") as f:
            json.dump(need, f)
        for name in ("alpha", "beta"):
            sol = os.path.join(d, "solutions", name)
            os.makedirs(sol)
            with open(os.path.join(sol, "features.json"), "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "solution_name": name,
                        "categories": [
                            {
                                "name": "Core",
                                "features": [
                                    {"name": "SSO", "supported": name == "alpha"},
                                    {"name": "API", "supported": True},
                                ],
                            }
                        ],
                    },
                    f,
                )

    @unittest.skipUnless(_has_openpyxl(), "openpyxl not installed")
    def test_compare_solutions_emits_xlsx_by_default(self) -> None:
        compare = os.path.join(SCRIPTS, "compare_solutions.py")
        with tempfile.TemporaryDirectory() as d:
            self._seed_compare_fixture(d)
            code, data = run_py(compare, "--dir", d)
            self.assertEqual(code, 0)
            self.assertEqual(data["status"], "ok")
            xlsx = str(comparison_matrix_xlsx_path(Path(d)))
            self.assertTrue(os.path.isfile(xlsx))
            self.assertIn("xlsx", data)

    @unittest.skipUnless(_has_openpyxl(), "openpyxl not installed")
    def test_generate_comparison_xlsx_from_existing_json(self) -> None:
        gen = os.path.join(SCRIPTS, "generate_comparison_xlsx.py")
        with tempfile.TemporaryDirectory() as d:
            self._seed_compare_fixture(d, persona_id="startup")
            compare = os.path.join(SCRIPTS, "compare_solutions.py")
            run_py(compare, "--dir", d, "--no-emit-xlsx")
            code, data = run_py(gen, "--dir", d)
            self.assertEqual(code, 0)
            self.assertEqual(data["status"], "ok")
            xlsx = data["output"]
            self.assertTrue(os.path.isfile(xlsx))

            import openpyxl

            wb = openpyxl.load_workbook(xlsx)
            ws = wb.active
            self.assertIn("startup", str(ws.cell(1, 1).value))
            self.assertEqual(ws.cell(2, 1).value, "Capability / Feature")
            self.assertEqual(ws.cell(2, 3).value, "alpha")
            self.assertEqual(ws.cell(5, 1).value, "Core")
            # SSO row: alpha tick, beta empty
            sso_row = None
            for row in range(5, ws.max_row + 1):
                if ws.cell(row, 1).value == "SSO":
                    sso_row = row
                    break
            self.assertIsNotNone(sso_row)
            assert sso_row is not None
            self.assertEqual(ws.cell(sso_row, 3).value, TICK)
            self.assertEqual(ws.cell(sso_row, 4).value, "\u2014")
            score_formula = str(ws.cell(4, 3).value or "")
            self.assertTrue(score_formula.startswith("="))
            self.assertIn("COUNTIFS", score_formula)
            self.assertIn(TICK, score_formula)
            self.assertNotIn('"?*"', score_formula)

    @unittest.skipUnless(_has_openpyxl(), "openpyxl not installed")
    def test_distinct_scores_for_different_feature_support(self) -> None:
        """Two solutions with different ticks must get different weighted scores."""
        from matrix_ops import ranked_scores

        compare = os.path.join(SCRIPTS, "compare_solutions.py")
        with tempfile.TemporaryDirectory() as d:
            self._seed_compare_fixture(d)
            code, data = run_py(compare, "--dir", d)
            self.assertEqual(code, 0)
            xlsx = str(comparison_matrix_xlsx_path(Path(d)))
            rs = ranked_scores(xlsx)
            by_plat = {r["platform"]: r["score"] for r in rs["rankings"]}
            self.assertNotEqual(by_plat["alpha"], by_plat["beta"])
            self.assertGreater(by_plat["alpha"], by_plat["beta"])

            comp = json.loads(
                Path(d, "comparison", "comparison.json").read_text(encoding="utf-8")
            )
            json_scores = {r["solution"]: r["score"] for r in comp["rankings"]}
            self.assertEqual(json_scores["alpha"], by_plat["alpha"])
            self.assertEqual(json_scores["beta"], by_plat["beta"])


class TestMatrixCoreNaming(unittest.TestCase):
    def test_comparison_matrix_xlsx_filename(self) -> None:
        from matrix_core import comparison_matrix_xlsx_filename

        self.assertEqual(
            comparison_matrix_xlsx_filename("2026-07-19-agentic-sdlc-orchestration-landscape-fullpool"),
            "2026-07-19-agentic-sdlc-orchestration-landscape-fullpool-comparison-matrix.xlsx",
        )

    def test_comparison_matrix_xlsx_path_uses_dir_slug(self) -> None:
        from matrix_core import comparison_matrix_xlsx_path, default_spa_report_path

        with tempfile.TemporaryDirectory(
            prefix="2026-07-19-agentic-sdlc-orchestration-landscape-fullpool-"
        ) as d:
            root = Path(d)
            slug = research_dir_slug(root)
            (root / "run_manifest.json").write_text(
                json.dumps({"research_type": "solution-landscape"}),
                encoding="utf-8",
            )
            self.assertEqual(
                default_spa_report_path(root).name,
                "landscape-report.html",
            )
            self.assertEqual(
                comparison_matrix_xlsx_path(root).name,
                f"{slug}-comparison-matrix.xlsx",
            )


class TestMatrixBuilderSanitization(unittest.TestCase):
    def test_user_strings_neutralize_formula_prefixes(self) -> None:
        from matrix_builder import _sanitize_cell_value

        self.assertEqual(_sanitize_cell_value("=SUM(A1)"), "'=SUM(A1)")
        self.assertEqual(_sanitize_cell_value("+cmd"), "'+cmd")
        self.assertEqual(_sanitize_cell_value("-1"), "'-1")
        self.assertEqual(_sanitize_cell_value("@evil"), "'@evil")
        self.assertEqual(_sanitize_cell_value("\t=SUM(A1)"), "'\t=SUM(A1)")
        self.assertEqual(_sanitize_cell_value("\r=1"), "'\r=1")
        self.assertEqual(_sanitize_cell_value(" Safe Name"), " Safe Name")


if __name__ == "__main__":
    unittest.main()

