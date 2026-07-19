#!/usr/bin/env python3
"""
Build a comparison matrix XLSX from scratch.

Produces a fully styled XLSX with:
  - Row 1: Title (merged across all columns)
  - Row 2: Column headers (Capability / Feature | Priority | Platform1 | ...)
  - Row 3: COUNTIF row (total capabilities per platform)
  - Row 4: Score row (priority-weighted COUNTIFS per platform)
  - Row 5+: Data rows (category headings + feature rows with ticks)
  - Freeze panes at A5
  - Auto-filter on header row

Input: a JSON config file:
{
  "title": "Capabilities Comparison Matrix",
  "categories": [
    {
      "name": "1. Category Name",
      "features": [
        {"name": "Feature A", "priority": "High"},
        {"name": "Feature B", "priority": "Critical"}
      ]
    }
  ],
  "platforms": [
    {"name": "Platform1", "features": ["Feature A", "Feature B"]},
    {"name": "Platform2", "features": ["Feature A"]}
  ]
}

When an existing matrix XLSX is provided via --clone-style, styles are
cloned from it. Otherwise, built-in defaults are used.

CLI:
    python3 matrix_builder.py --config build.json --out matrix.xlsx [--clone-style existing.xlsx]
"""
from __future__ import annotations

import argparse
import copy
import json
import logging
import sys
from pathlib import Path
from typing import Any, Optional

from matrix_core import TICK, TICK_CRITERION, WEIGHTS

openpyxl = None
get_column_letter = None


def _require_openpyxl() -> None:
    global openpyxl, get_column_letter
    if openpyxl is not None:
        return
    try:
        import openpyxl as _openpyxl
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter as _gcl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for matrix_builder. "
            "Install with: pip install -r skills/silver-deep-research/requirements.txt"
        ) from exc
    openpyxl = _openpyxl
    get_column_letter = _gcl
    globals().update({
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
    })
    _init_styles()


Worksheet = Any

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Layout constants (new matrices always use the "with-title" layout)
# ---------------------------------------------------------------------------
TITLE_ROW = 1
HEADER_ROW = 2
TOTAL_ROW = 3
SCORE_ROW = 4
DATA_START = 5
FEAT_COL = 1
PRIO_COL = 2
PLAT_START = 3

# TICK and WEIGHTS from matrix_core

# Default styles — initialized lazily via _require_openpyxl()
_THIN = None
_DEFAULT_BORDER = None
_TITLE_FONT = None
_TITLE_FILL = None
_TITLE_ALIGN = None
_HEADER_FONT = None
_HEADER_FILL = None
_HEADER_ALIGN = None
_TOTAL_FONT = None
_TOTAL_FILL = None
_TOTAL_ALIGN = None
_SCORE_FONT = None
_SCORE_FILL = None
_SCORE_ALIGN = None
_CAT_FONT = None
_CAT_FILL = None
_CAT_ALIGN = None
_FEAT_FONT = None
_FEAT_ALIGN = None
_TICK_FONT = None
_TICK_FILL = None
_TICK_ALIGN = None
_EMPTY_FILL = None
PRIORITY_FILLS: dict[str, Any] = {}
_PRIO_FONT = None
_PRIO_ALIGN = None


_FONT_FAMILY = "Roboto Condensed"


def _init_styles() -> None:
    global _THIN, _DEFAULT_BORDER, _TITLE_FONT, _TITLE_FILL, _TITLE_ALIGN
    global _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN
    global _TOTAL_FONT, _TOTAL_FILL, _TOTAL_ALIGN
    global _SCORE_FONT, _SCORE_FILL, _SCORE_ALIGN
    global _CAT_FONT, _CAT_FILL, _CAT_ALIGN
    global _FEAT_FONT, _FEAT_ALIGN
    global _TICK_FONT, _TICK_FILL, _TICK_ALIGN
    global _EMPTY_FILL, PRIORITY_FILLS, _PRIO_FONT, _PRIO_ALIGN
    if _THIN is not None:
        return
    _THIN = Side(style="thin", color="D0D7DE")
    _DEFAULT_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    # Restrained professional blues/grays (Roboto Condensed — system may fall back if absent)
    _TITLE_FONT = Font(name=_FONT_FAMILY, bold=True, size=14, color="FFFFFF")
    _TITLE_FILL = PatternFill("solid", fgColor="1F3864")
    _TITLE_ALIGN = Alignment(horizontal="center", vertical="center")
    _HEADER_FONT = Font(name=_FONT_FAMILY, bold=True, size=11, color="FFFFFF")
    _HEADER_FILL = PatternFill("solid", fgColor="2F5597")
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _TOTAL_FONT = Font(name=_FONT_FAMILY, bold=True, size=11, color="1F3864")
    _TOTAL_FILL = PatternFill("solid", fgColor="E8EEF4")
    _TOTAL_ALIGN = Alignment(horizontal="center", vertical="center")
    _SCORE_FONT = Font(name=_FONT_FAMILY, bold=True, size=11, color="1F3864")
    _SCORE_FILL = PatternFill("solid", fgColor="E8EEF4")
    _SCORE_ALIGN = Alignment(horizontal="center", vertical="center")
    _CAT_FONT = Font(name=_FONT_FAMILY, bold=True, size=11, color="FFFFFF")
    _CAT_FILL = PatternFill("solid", fgColor="5B7FA6")
    _CAT_ALIGN = Alignment(horizontal="left", vertical="center")
    _FEAT_FONT = Font(name=_FONT_FAMILY, size=10, color="24292F")
    _FEAT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # Tick cells: neutral black on white (no green fills)
    _TICK_FONT = Font(name=_FONT_FAMILY, size=10, color="1F1F1F")
    _TICK_FILL = PatternFill("solid", fgColor="FFFFFF")
    _TICK_ALIGN = Alignment(horizontal="center", vertical="center")
    _EMPTY_FILL = PatternFill("solid", fgColor="FFFFFF")
    PRIORITY_FILLS = {
        "Critical": PatternFill("solid", fgColor="F4C7C3"),
        "Very High": PatternFill("solid", fgColor="F9DCC4"),
        "High": PatternFill("solid", fgColor="FCE8B2"),
        "Medium": PatternFill("solid", fgColor="E8EEF4"),
        "Low": PatternFill("solid", fgColor="F3F4F6"),
    }
    _PRIO_FONT = Font(name=_FONT_FAMILY, size=10, color="24292F")
    _PRIO_ALIGN = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _sanitize_cell_value(value: Any) -> Any:
    """Neutralize spreadsheet formula injection for user-controlled strings."""
    if not isinstance(value, str) or not value:
        return value
    lead = value.lstrip("\t\r\n ")
    if lead and lead[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _countif(col_letter: str) -> str:
    return f'=COUNTIF({col_letter}{DATA_START}:{col_letter}1048576,"{TICK_CRITERION}")'


def _score_formula(col_letter: str) -> str:
    b = f"$B${DATA_START}:$B$1048576"
    c = f"{col_letter}${DATA_START}:{col_letter}$1048576"
    parts = []
    for prio, wt in WEIGHTS.items():
        parts.append(f'COUNTIFS({b},"{prio}",{c},"{TICK_CRITERION}")*{wt}')
    return "=" + "+".join(parts)


def _style_cell(cell, font=None, fill=None, align=None, border=None):
    """Apply style to a cell (only non-None attributes)."""
    if font:
        cell.font = copy.copy(font)
    if fill:
        cell.fill = copy.copy(fill)
    if align:
        cell.alignment = copy.copy(align)
    if border:
        cell.border = copy.copy(border)


def _set_user_cell(ws, row: int, col: int, value: Any, font=None, fill=None, align=None, border=None):
    """Write a user-controlled value with formula neutralization + optional style."""
    cell = ws.cell(row, col)
    cell.value = _sanitize_cell_value(value)
    _style_cell(cell, font, fill, align, border)
    return cell


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------

def build_matrix(config: dict, out_xlsx: str, clone_xlsx: Optional[str] = None) -> dict:
    """Build a comparison matrix from JSON config.

    Args:
        config: {title, categories: [{name, features: [{name, priority}]}],
                 platforms: [{name, features: [feat_name, ...]}]}
        out_xlsx: Output path.
        clone_xlsx: Optional existing XLSX to clone styles from.

    Returns:
        {platforms_added, categories, features, total_rows}
    """
    _require_openpyxl()
    if clone_xlsx is not None:
        log.warning(
            "--clone-style / clone_xlsx is not yet implemented; "
            "the new matrix will use default openpyxl styles."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Matrix"

    title = config.get("title", "Capabilities Comparison Matrix")
    categories = config.get("categories", [])
    platforms = config.get("platforms", [])

    num_plats = len(platforms)
    last_col = PLAT_START + num_plats - 1
    last_ltr = get_column_letter(last_col)

    # ------------------------------------------------------------------
    # Row 1: Title
    # ------------------------------------------------------------------
    ws.merge_cells(f"A{TITLE_ROW}:{last_ltr}{TITLE_ROW}")
    _set_user_cell(
        ws, TITLE_ROW, 1, title,
        _TITLE_FONT, _TITLE_FILL, _TITLE_ALIGN, _DEFAULT_BORDER,
    )
    ws.row_dimensions[TITLE_ROW].height = 30

    # ------------------------------------------------------------------
    # Row 2: Headers
    # ------------------------------------------------------------------
    h_cap = ws.cell(HEADER_ROW, FEAT_COL)
    h_cap.value = "Capability / Feature"
    _style_cell(h_cap, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _DEFAULT_BORDER)

    h_prio = ws.cell(HEADER_ROW, PRIO_COL)
    h_prio.value = "Priority"
    _style_cell(h_prio, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _DEFAULT_BORDER)

    for p_idx, plat in enumerate(platforms):
        col = PLAT_START + p_idx
        _set_user_cell(
            ws, HEADER_ROW, col, plat["name"],
            _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, _DEFAULT_BORDER,
        )

    ws.row_dimensions[HEADER_ROW].height = 30

    # ------------------------------------------------------------------
    # Row 3: COUNTIF (Total Capabilities)
    # ------------------------------------------------------------------
    tc = ws.cell(TOTAL_ROW, FEAT_COL)
    tc.value = "Total Capabilities"
    _style_cell(tc, _TOTAL_FONT, _TOTAL_FILL, Alignment(horizontal="left", vertical="center"), _DEFAULT_BORDER)

    tc_p = ws.cell(TOTAL_ROW, PRIO_COL)
    tc_p.value = "\u2014"
    _style_cell(tc_p, _TOTAL_FONT, _TOTAL_FILL, _TOTAL_ALIGN, _DEFAULT_BORDER)

    for p_idx in range(num_plats):
        col = PLAT_START + p_idx
        ltr = get_column_letter(col)
        cell = ws.cell(TOTAL_ROW, col)
        cell.value = _countif(ltr)
        _style_cell(cell, _TOTAL_FONT, _TOTAL_FILL, _TOTAL_ALIGN, _DEFAULT_BORDER)

    # ------------------------------------------------------------------
    # Row 4: Score (Weighted)
    # ------------------------------------------------------------------
    sc = ws.cell(SCORE_ROW, FEAT_COL)
    sc.value = "Score"
    _style_cell(sc, _SCORE_FONT, _SCORE_FILL, Alignment(horizontal="left", vertical="center"), _DEFAULT_BORDER)

    sc_p = ws.cell(SCORE_ROW, PRIO_COL)
    sc_p.value = "\u2014"
    _style_cell(sc_p, _SCORE_FONT, _SCORE_FILL, _SCORE_ALIGN, _DEFAULT_BORDER)

    for p_idx in range(num_plats):
        col = PLAT_START + p_idx
        ltr = get_column_letter(col)
        cell = ws.cell(SCORE_ROW, col)
        cell.value = _score_formula(ltr)
        _style_cell(cell, _SCORE_FONT, _SCORE_FILL, _SCORE_ALIGN, _DEFAULT_BORDER)

    # ------------------------------------------------------------------
    # Row 5+: Data (categories + features)
    # ------------------------------------------------------------------
    # Build per-platform feature sets for fast lookup
    plat_feats: list[set[str]] = []
    for plat in platforms:
        plat_feats.append(set(plat.get("features", [])))

    current_row = DATA_START
    total_features = 0

    for cat in categories:
        # Category heading row (merged)
        ws.merge_cells(f"A{current_row}:{last_ltr}{current_row}")
        _set_user_cell(
            ws, current_row, FEAT_COL, cat["name"],
            _CAT_FONT, _CAT_FILL, _CAT_ALIGN, _DEFAULT_BORDER,
        )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for feat in cat.get("features", []):
            feat_name = feat["name"]
            prio = feat.get("priority", "Medium")
            total_features += 1

            # Col A: feature name
            _set_user_cell(
                ws, current_row, FEAT_COL, feat_name,
                _FEAT_FONT, _EMPTY_FILL, _FEAT_ALIGN, _DEFAULT_BORDER,
            )

            # Col B: priority
            prio_fill = PRIORITY_FILLS.get(prio, _EMPTY_FILL)
            _set_user_cell(
                ws, current_row, PRIO_COL, prio,
                _PRIO_FONT, prio_fill, _PRIO_ALIGN, _DEFAULT_BORDER,
            )

            # Platform columns: tick or empty
            for p_idx in range(num_plats):
                col = PLAT_START + p_idx
                cell = ws.cell(current_row, col)
                cell.border = copy.copy(_DEFAULT_BORDER)
                if feat_name in plat_feats[p_idx]:
                    cell.value = TICK
                    _style_cell(cell, _TICK_FONT, _TICK_FILL, _TICK_ALIGN, _DEFAULT_BORDER)
                else:
                    cell.value = "\u2014"
                    _style_cell(cell, _FEAT_FONT, _EMPTY_FILL, _TICK_ALIGN, _DEFAULT_BORDER)

            current_row += 1

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    for p_idx in range(num_plats):
        ltr = get_column_letter(PLAT_START + p_idx)
        ws.column_dimensions[ltr].width = 18

    # ------------------------------------------------------------------
    # Freeze panes & auto-filter
    # ------------------------------------------------------------------
    ws.freeze_panes = f"A{DATA_START}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_ltr}{HEADER_ROW}"
    ws.sheet_view.showGridLines = False

    wb.save(out_xlsx)

    return {
        "platforms_added": num_plats,
        "platform_names": [p["name"] for p in platforms],
        "categories": len(categories),
        "features": total_features,
        "total_rows": current_row - 1,
        "output": out_xlsx,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    _require_openpyxl()
    parser = argparse.ArgumentParser(
        description="Build a comparison matrix XLSX from JSON config",
    )
    parser.add_argument("--config", required=True,
                        help="JSON config file with title, categories, platforms")
    parser.add_argument("--out", required=True,
                        help="Output XLSX path")
    parser.add_argument("--clone-style", default=None,
                        help="Optional existing XLSX to clone styles from (reserved)")
    args = parser.parse_args()

    config = json.loads(Path(args.config).read_text(encoding="utf-8"))
    result = build_matrix(config, args.out, clone_xlsx=args.clone_style)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()


